#!/usr/bin/env python3
"""
🏠 분양 아파트 만료 도메인 헌터 v3.4
=====================================
구매가능일(낙장일) 정확 계산:

  .kr / .co.kr:
    만료 → 유예30일 → 삭제 (오전 9시)
    구매가능일 = 만료일 + 31일

  .com / .net (국제 도메인) — WHOIS 상태 기반 계산:
    ① pendingDelete 상태 + Updated Date → Updated Date + 5일 (★ 가장 정확)
    ② redemptionPeriod 상태 + Updated Date → Updated Date + 35일 (★ 정확)
    ③ 상태 모름 (만료일만 있음) → 만료일 + 35~80일 (범위 추정)

    삭제 시간: KST 새벽 3~5시 (미국 서부시간 11AM~2PM)
"""

import socket, subprocess, json, os, sys, time, re, argparse, ssl
import concurrent.futures
from datetime import datetime, timedelta
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError
from urllib.parse import urlencode, urlparse, quote

ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE

# ============================================================
# 1. 청약홈 API
# ============================================================

API_ENDPOINTS = {
    "APT": {"name": "APT 분양정보", "url": "https://api.odcloud.kr/api/ApplyhomeInfoDetailSvc/v1/getAPTLttotPblancDetail"},
}

def fetch_applyhome(service_key, max_pages=50):
    ep = API_ENDPOINTS["APT"]
    all_data, page = [], 1
    print(f"\n📡 [{ep['name']}] 수집 중...")
    while page <= max_pages:
        params = {"page": page, "perPage": 500, "serviceKey": service_key}
        url = f"{ep['url']}?{urlencode(params, quote_via=quote)}"
        try:
            req = Request(url, headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0"})
            data = json.loads(urlopen(req, timeout=30).read().decode("utf-8"))
            if "data" not in data: break
            items = data["data"]; total = data.get("totalCount", 0)
            if not items: break
            all_data.extend(items)
            print(f"   📄 페이지 {page}: {len(items)}건 (누적 {len(all_data)}/{total})")
            if len(all_data) >= total: break
            page += 1; time.sleep(0.3)
        except Exception as e:
            print(f"   ❌ {e}"); break
    print(f"   ✅ 총 {len(all_data)}건")
    return all_data

def filter_by_date(api_data, date_from=None, date_to=None):
    if not date_from and not date_to: return api_data
    return [i for i in api_data
            if (i.get("RCRIT_PBLANC_DE") or "") and
               (not date_from or (i.get("RCRIT_PBLANC_DE") or "") >= date_from) and
               (not date_to or (i.get("RCRIT_PBLANC_DE") or "") <= date_to)]

def extract_domains(api_data):
    complexes, seen = [], set()
    for item in api_data:
        name = (item.get("HOUSE_NM") or "").strip()
        homepage = (item.get("HMPG_ADRES") or "").strip()
        if not name: continue
        domains = []
        if homepage:
            for u in re.split(r'[,\s]+', homepage):
                u = u.strip()
                if not u: continue
                if not u.startswith("http"): u = "http://" + u
                try:
                    dom = urlparse(u).netloc.lower().replace("www.", "")
                    if dom and dom not in seen: domains.append(dom); seen.add(dom)
                except: pass
        complexes.append({"name": name, "homepage": homepage, "domains": domains,
                          "region": (item.get("SUBSCRPT_AREA_CODE_NM") or ""),
                          "notice_date": (item.get("RCRIT_PBLANC_DE") or "")})
    return complexes


# ============================================================
# 2. WHOIS (상태 + Updated Date 파싱 추가)
# ============================================================

WHOIS_SERVERS = {
    "kr": "whois.kr", "co.kr": "whois.kr", "or.kr": "whois.kr",
    "ne.kr": "whois.kr", "pe.kr": "whois.kr", "go.kr": "whois.kr",
    "com": "whois.verisign-grs.com", "net": "whois.verisign-grs.com",
    "org": "whois.pir.org",
}
SKIP_DOMAINS = ["naver.com","daum.net","tistory.com","modoo.at","imweb.me",
                "wixsite.com","iwinv.net","quv.kr","kro.kr"]

def to_punycode(domain):
    try: return ".".join(p.encode("idna").decode("ascii") for p in domain.split("."))
    except: return domain

def get_main_domain(domain):
    domain = domain.split(":")[0].lower().strip()
    parts = domain.split(".")
    kr_slds = ["co","or","ne","go","pe","re","ac"]
    if len(parts) >= 3 and parts[-1] == "kr" and parts[-2] in kr_slds:
        return ".".join(parts[-3:])
    if len(parts) >= 2: return ".".join(parts[-2:])
    return domain

def get_tld(domain):
    parts = domain.split(".")
    kr_slds = ["co","or","ne","go","pe","re","ac"]
    if len(parts) >= 2 and parts[-1] == "kr" and parts[-2] in kr_slds:
        return f"{parts[-2]}.{parts[-1]}"
    return parts[-1] if parts else ""

def whois_socket(domain, server, timeout=10):
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(timeout)
        s.connect((server, 43))
        s.send(f"{domain}\r\n".encode())
        result = b""
        while True:
            data = s.recv(4096)
            if not data: break
            result += data
        s.close()
        return result.decode("utf-8", errors="ignore")
    except: return ""

def whois_command(domain, timeout=10):
    try:
        r = subprocess.run(["whois", domain], capture_output=True, text=True, timeout=timeout)
        return r.stdout
    except: return ""

def parse_date_str(raw):
    """날짜 문자열을 YYYY-MM-DD로 정규화"""
    if not raw: return ""
    raw = raw.strip()
    m = re.match(r"(\d{4}-\d{2}-\d{2})", raw)
    if m: return m.group(1)
    m = re.match(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})", raw)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return ""

def parse_time_str(raw):
    """날짜 문자열에서 시간 부분 추출"""
    if not raw: return ""
    m = re.search(r"(\d{2}:\d{2}:\d{2})", raw)
    return m.group(1) if m else ""

def parse_whois(text):
    """WHOIS 텍스트에서 만료일 + 상태 + Updated Date 추출"""
    info = {"만료일_원본": "", "등록일": "", "상태": "", "updated_date": ""}
    if not text: return info
    tl = text.lower()

    # 미등록
    not_found = ["no match for","not found","no data found","is free","available",
                  "no entries found","도메인이름이 등록되어 있지 않습니다",
                  "above domain name is not registered","this query returned 0 objects"]
    for nf in not_found:
        if nf in tl:
            info["만료일_원본"] = "🎯 미등록"; return info

    # 만료일
    for p in [r"Registry Expiry Date:\s*(.+)", r"Registrar Registration Expiration Date:\s*(.+)",
              r"Expiration Date\s*:\s*(.+)", r"Expiry Date\s*:\s*(.+)",
              r"등록기간만료일\s*:\s*(.+)", r"expire\s*:\s*(.+)"]:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if len(val) >= 8: info["만료일_원본"] = val; break

    # 등록일
    for p in [r"Creation Date:\s*(.+)", r"등록일\s*:\s*(.+)"]:
        m = re.search(p, text, re.IGNORECASE)
        if m: info["등록일"] = m.group(1).strip(); break

    # Updated Date (낙장일 계산의 핵심!)
    for p in [r"Updated Date:\s*(.+)", r"최근변경일\s*:\s*(.+)"]:
        m = re.search(p, text, re.IGNORECASE)
        if m: info["updated_date"] = m.group(1).strip(); break

    # Domain Status (redemptionPeriod, pendingDelete 등)
    statuses = re.findall(r"(?:Domain )?Status:\s*(\S+)", text, re.IGNORECASE)
    status_list = [s.lower() for s in statuses]

    if any("pendingdelete" in s for s in status_list):
        info["상태"] = "pendingDelete"
    elif any("redemption" in s for s in status_list):
        info["상태"] = "redemptionPeriod"
    elif any("hold" in s or "expired" in s for s in status_list):
        info["상태"] = "expired/hold"
    elif any("ok" in s or "active" in s for s in status_list):
        info["상태"] = "active"

    return info


def calc_drop_date(expiry_str, updated_str, status, tld):
    """
    낙장일(구매가능일) 계산 — WHOIS 상태 기반

    .kr: 만료일 + 31일 (고정)
    .com/.net (상태별):
      pendingDelete + Updated Date → Updated Date + 5일
      redemptionPeriod + Updated Date → Updated Date + 35일
      상태 모름 → 만료일 + 35~80일 (범위)
    """
    result = {"구매가능일": "", "구매가능일_범위": "", "계산근거": "", "삭제시간": ""}

    # 미등록
    if "미등록" in (expiry_str or ""):
        result["구매가능일"] = "즉시 등록 가능"
        result["계산근거"] = "WHOIS 미등록"
        return result

    is_kr = tld in ("kr","co.kr","or.kr","ne.kr","pe.kr","go.kr")
    exp_date = parse_date_str(expiry_str)
    upd_date = parse_date_str(updated_str)

    if is_kr:
        # .kr: 만료일 + 31일 (유예30일 → 다음날 오전9시 삭제)
        if exp_date:
            try:
                d = datetime.strptime(exp_date, "%Y-%m-%d") + timedelta(days=31)
                result["구매가능일"] = d.strftime("%Y-%m-%d")
                result["계산근거"] = "만료일+31일"
                result["삭제시간"] = "오전 9:00~10:00 KST"
            except: pass
        return result

    # 국제 도메인 (.com/.net/.org 등)
    result["삭제시간"] = "새벽 3:00~5:00 KST"

    if status == "pendingDelete" and upd_date:
        # ★★★ 가장 정확: Updated Date + 5일 (다음날 새벽)
        try:
            d = datetime.strptime(upd_date, "%Y-%m-%d") + timedelta(days=5)
            result["구매가능일"] = d.strftime("%Y-%m-%d")
            result["계산근거"] = f"pendingDelete→Updated({upd_date})+5일"
            return result
        except: pass

    if status == "redemptionPeriod" and upd_date:
        # ★★☆ 정확: Updated Date + 30일(복구) + 5일(PD)
        try:
            d = datetime.strptime(upd_date, "%Y-%m-%d") + timedelta(days=35)
            result["구매가능일"] = d.strftime("%Y-%m-%d")
            result["계산근거"] = f"redemption→Updated({upd_date})+35일"
            return result
        except: pass

    # ★☆☆ 추정: 만료일 기준 범위
    if exp_date:
        try:
            exp = datetime.strptime(exp_date, "%Y-%m-%d")
            earliest = (exp + timedelta(days=35)).strftime("%Y-%m-%d")
            typical = (exp + timedelta(days=75)).strftime("%Y-%m-%d")
            latest = (exp + timedelta(days=80)).strftime("%Y-%m-%d")
            result["구매가능일"] = typical
            result["구매가능일_범위"] = f"{earliest} ~ {latest}"
            result["계산근거"] = "만료일+35~80일 (상태 미확인, 추정)"
        except: pass

    return result


def check_whois(domain):
    main_dom = get_main_domain(domain)
    result = {"만료일": "", "만료시간": "", "등록일": "", "도메인상태": "", "Updated Date": "",
              "구매가능일": "", "구매가능일_범위": "", "계산근거": "", "삭제시간": ""}
    if any(main_dom == s or domain.endswith("." + s) for s in SKIP_DOMAINS): return result

    tld = get_tld(main_dom)
    query_dom = to_punycode(main_dom)

    # 1순위: 소켓
    server = WHOIS_SERVERS.get(tld)
    if server:
        raw = whois_socket(query_dom, server)
        if raw:
            info = parse_whois(raw)
            if info["만료일_원본"]:
                result["만료일"] = parse_date_str(info["만료일_원본"])
                result["만료시간"] = parse_time_str(info["만료일_원본"])
                result["등록일"] = info["등록일"]
                result["도메인상태"] = info["상태"]
                result["Updated Date"] = parse_date_str(info["updated_date"])
                drop = calc_drop_date(info["만료일_원본"], info["updated_date"], info["상태"], tld)
                result.update(drop)
                return result

    # 2순위: 명령줄 whois
    for try_dom in [query_dom, main_dom]:
        raw = whois_command(try_dom)
        if raw:
            info = parse_whois(raw)
            if info["만료일_원본"]:
                result["만료일"] = parse_date_str(info["만료일_원본"])
                result["만료시간"] = parse_time_str(info["만료일_원본"])
                result["등록일"] = info["등록일"]
                result["도메인상태"] = info["상태"]
                result["Updated Date"] = parse_date_str(info["updated_date"])
                drop = calc_drop_date(info["만료일_원본"], info["updated_date"], info["상태"], tld)
                result.update(drop)
                return result
    return result


# ============================================================
# 3. 체크 + Excel
# ============================================================

def check_domain(domain):
    r = {"도메인": domain, "만료일": "", "만료시간": "", "등록일": "", "도메인상태": "",
         "Updated Date": "", "구매가능일": "", "구매가능일_범위": "", "계산근거": "", "삭제시간": ""}
    w = check_whois(domain)
    r.update(w)
    time.sleep(0.5)
    return r

def try_date(s):
    try: return datetime.strptime(s, "%Y-%m-%d")
    except: return None

def save_excel(results, complexes, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    brd = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    hfill = PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
    hfont = Font(color="FFFFFF",bold=True,size=11)
    today = datetime.now()

    ws = wb.active; ws.title = "전체 결과"
    hdrs = ["지역","단지명","공고일","원본URL","도메인","구매가능일","만료일","등록일"]
    for c,h in enumerate(hdrs,1):
        cl=ws.cell(row=1,column=c,value=h); cl.fill=hfill; cl.font=hfont
        cl.alignment=Alignment(horizontal="center"); cl.border=brd

    ncols = len(hdrs)
    for i,item in enumerate(results,2):
        # 구매가능일: 날짜 + 삭제시간
        drop = item["구매가능일"]
        if drop and "즉시" not in drop and item.get("삭제시간"):
            drop += f" ({item['삭제시간']})"

        # 만료일 + 시간
        exp = item["만료일"]
        if exp and item.get("만료시간"):
            exp += f" {item['만료시간']}"

        vals = [item.get("_region",""), item.get("_name",""), item.get("_notice_date",""),
                item.get("_homepage",""), item["도메인"],
                drop, exp, item["등록일"]]
        for c,v in enumerate(vals,1): ws.cell(row=i,column=c,value=v).border=brd

        p = item.get("구매가능일","")
        if "즉시" in p:
            f=PatternFill(start_color="92D050",end_color="92D050",fill_type="solid")
            for c in range(1,ncols+1): ws.cell(row=i,column=c).fill=f
        elif p:
            pd = try_date(p)
            if pd:
                diff = (pd - today).days
                color = None
                if diff <= 0: color = "92D050"
                elif diff <= 30: color = "FFC000"
                elif diff <= 90: color = "FFFF00"
                if color:
                    f=PatternFill(start_color=color,end_color=color,fill_type="solid")
                    for c in range(1,ncols+1): ws.cell(row=i,column=c).fill=f

    widths = {"A":10,"B":30,"C":12,"D":45,"E":35,"F":28,"G":24,"H":24}
    for col,w in widths.items(): ws.column_dimensions[col].width=w

    # 시트2: 지금 구매 가능
    ws2 = wb.create_sheet("⭐ 지금 구매 가능")
    h2 = ["지역","단지명","공고일","원본URL","도메인","구매가능일","만료일","등록일"]
    for c,h in enumerate(h2,1):
        cl=ws2.cell(row=1,column=c,value=h)
        cl.fill=PatternFill(start_color="00B050",end_color="00B050",fill_type="solid")
        cl.font=Font(color="FFFFFF",bold=True)
    row=2
    for r in results:
        p = r.get("구매가능일","")
        avail = "즉시" in p or (p and try_date(p) and try_date(p) <= today)
        if avail:
            drop = p
            if drop and "즉시" not in drop and r.get("삭제시간"): drop += f" ({r['삭제시간']})"
            exp = r["만료일"]
            if exp and r.get("만료시간"): exp += f" {r['만료시간']}"
            for c,v in enumerate([r.get("_region",""),r.get("_name",""),r.get("_notice_date",""),
                                   r.get("_homepage",""),r["도메인"],drop,exp,r["등록일"]],1):
                ws2.cell(row=row,column=c,value=v)
            row+=1

    # 시트3: 30일 이내
    ws3 = wb.create_sheet("🔥 30일 이내")
    h3 = ["지역","단지명","공고일","원본URL","도메인","구매가능일","만료일","등록일"]
    for c,h in enumerate(h3,1):
        cl=ws3.cell(row=1,column=c,value=h)
        cl.fill=PatternFill(start_color="FF6600",end_color="FF6600",fill_type="solid")
        cl.font=Font(color="FFFFFF",bold=True)
    row=2
    for r in results:
        p = r.get("구매가능일","")
        if p and "즉시" not in p:
            pd = try_date(p)
            if pd and 0 < (pd-today).days <= 30:
                drop = p
                if r.get("삭제시간"): drop += f" ({r['삭제시간']})"
                exp = r["만료일"]
                if exp and r.get("만료시간"): exp += f" {r['만료시간']}"
                for c,v in enumerate([r.get("_region",""),r.get("_name",""),r.get("_notice_date",""),
                                       r.get("_homepage",""),r["도메인"],drop,exp,r["등록일"]],1):
                    ws3.cell(row=row,column=c,value=v)
                row+=1

    # 시트4: 요약
    ws4 = wb.create_sheet("요약")
    ws4.cell(row=1,column=1,value="🏠 도메인 헌터 v3.4 결과 요약").font=Font(bold=True,size=14)
    ws4.cell(row=2,column=1,value=f"조회일시: {today.strftime('%Y-%m-%d %H:%M:%S')}")
    ws4.cell(row=3,column=1,value=f"총 단지: {len(complexes)}개 / 도메인: {len(results)}개")
    ws4.cell(row=5,column=1,value="📋 구매가능일(낙장일) 계산 기준:").font=Font(bold=True)
    ws4.cell(row=6,column=1,value="  .kr: 만료일 + 31일 (유예30일 → 오전9시 삭제)")
    ws4.cell(row=7,column=1,value="  .com/.net (pendingDelete 상태): Updated Date + 5일 → 새벽3~5시 삭제")
    ws4.cell(row=8,column=1,value="  .com/.net (redemptionPeriod 상태): Updated Date + 35일 → 새벽3~5시 삭제")
    ws4.cell(row=9,column=1,value="  .com/.net (상태 미확인): 만료일 + 35~80일 (범위 추정)")
    ws4.column_dimensions["A"].width=65
    wb.save(filename)


# ============================================================
# 4. 메인
# ============================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--key", type=str, required=True)
    parser.add_argument("--date-from", type=str, default="")
    parser.add_argument("--date-to", type=str, default="")
    parser.add_argument("--output", type=str, default="domain_hunter_results")
    parser.add_argument("--workers", type=int, default=6)
    args = parser.parse_args()

    print("""
╔══════════════════════════════════════════════════════════════════╗
║   🏠 분양 아파트 만료 도메인 헌터 v3.4                           ║
║   ────────────────────────────────────────────────────────     ║
║   .kr: 만료+31일 / .com: WHOIS상태+Updated Date 기반 정밀 계산  ║
║   pendingDelete → +5일 / redemption → +35일 / 추정 → +35~80일  ║
╚══════════════════════════════════════════════════════════════════╝
""")

    all_data = fetch_applyhome(args.key)
    if args.date_from or args.date_to:
        all_data = filter_by_date(all_data, args.date_from, args.date_to)

    complexes = extract_domains(all_data)
    to_check = []
    for c in complexes:
        for d in c.get("domains", []):
            to_check.append({"domain": d, "name": c["name"], "homepage": c.get("homepage",""),
                             "region": c.get("region",""), "notice_date": c.get("notice_date","")})

    print(f"\n🔍 {len(to_check)}개 도메인 WHOIS 체크 중...\n")

    results, checked = [], 0
    total = len(to_check); t0 = time.time()

    def proc(item):
        r = check_domain(item["domain"])
        r["_name"]=item["name"]; r["_homepage"]=item["homepage"]
        r["_region"]=item["region"]; r["_notice_date"]=item["notice_date"]
        return r

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        fmap = {ex.submit(proc, i): i for i in to_check}
        for f in concurrent.futures.as_completed(fmap):
            checked+=1; r=f.result(); results.append(r)
            p=r.get("구매가능일",""); s=r.get("도메인상태","")
            w = ""
            if "즉시" in p: w=f" | ⭐ 즉시등록가능"
            elif p: w=f" | 구매: {p}"
            if s: w += f" [{s}]"
            if checked % 20 == 0 or "즉시" in p or s in ("pendingDelete","redemptionPeriod"):
                print(f"  [{checked:4d}/{total}] {r['도메인'][:30]:30s}{w}")

    results.sort(key=lambda r: r.get("구매가능일","") if r.get("구매가능일","") and "즉시" not in r.get("구매가능일","") else ("0000" if "즉시" in r.get("구매가능일","") else "9999"))

    print(f"\n⏱️  {time.time()-t0:.0f}초 / {len(results)}개 완료")

    save_excel(results, complexes, f"{args.output}.xlsx")
    print(f"💾 저장: {args.output}.xlsx")

    now_avail = [r for r in results if "즉시" in r.get("구매가능일","") or
                 (try_date(r.get("구매가능일","")) and try_date(r.get("구매가능일","")) <= datetime.now())]
    pending = [r for r in results if r.get("도메인상태") in ("pendingDelete","redemptionPeriod")]

    if now_avail:
        print(f"\n⭐ 지금 구매 가능: {len(now_avail)}개")
        for r in now_avail[:20]:
            print(f"  ⭐ {r['도메인']:35s} | {r.get('_name','')[:20]} | {r['계산근거']}")
    if pending:
        print(f"\n🚨 삭제 임박 (pendingDelete/redemption): {len(pending)}개")
        for r in pending[:20]:
            print(f"  🚨 {r['도메인']:35s} | {r['도메인상태']} | 구매: {r['구매가능일']}")

    print("\n✅ 완료!")

if __name__ == "__main__":
    main()
